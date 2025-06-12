import pandas as pd
import os
import time
import re
import openpyxl
# Import openpyxl.styles for copying styles
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.hyperlink import Hyperlink # Import Hyperlink object

def process_data_excel(input_filepath, cleaned_output_filepath, excluded_output_filepath, keywords_list=None, input_sheet_name='Sheet1', output_sheet_name='Processed Data'):

    if keywords_list is None:
        keywords_list = []

    original_workbook = None
    try:
        # Load the original workbook using openpyxl
        original_workbook = openpyxl.load_workbook(input_filepath)

        # Check if all sheets are hidden. If so, unhide the input_sheet_name temporarily.
        all_sheets_hidden = all(sheet.sheet_state == 'hidden' or sheet.sheet_state == 'veryHidden' for sheet in original_workbook._sheets)

        if all_sheets_hidden:
            if input_sheet_name not in original_workbook.sheetnames:
                raise ValueError(f"Input sheet '{input_sheet_name}' not found in the Excel file, and all sheets are hidden.")

            # Temporarily unhide the input sheet for openpyxl to access it
            original_workbook[input_sheet_name].sheet_state = 'visible'
            print(f"Temporarily unhid sheet '{input_sheet_name}' as all sheets were hidden.")

        # Get the input sheet for processing
        try:
            input_sheet = original_workbook[input_sheet_name]
        except KeyError:
            raise ValueError(f"Input sheet '{input_sheet_name}' not found in the Excel file.")

        # --- Identify 'KONTEN' and 'UUID' columns dynamically from the first row ---
        header_row_cells = [cell for cell in input_sheet[1]] # Get the actual cell objects for the header row
        konten_col_index = -1
        uuid_col_index = -1 # New: Index for the 'UUID' column
        
        for idx, cell in enumerate(header_row_cells):
            col_val = str(cell.value or '').strip().upper()
            if col_val == 'KONTEN':
                konten_col_index = idx # 0-indexed column
            elif col_val == 'UUID': # Identify 'UUID' column
                uuid_col_index = idx # 0-indexed column

        if konten_col_index == -1:
            raise ValueError(
                f"'KONTEN' column not found in sheet '{input_sheet_name}'. "
                f"Available columns (from first row) are: {[cell.value for cell in header_row_cells]}"
            )

        # Create new workbooks for cleaned and excluded data
        cleaned_workbook = openpyxl.Workbook()
        excluded_workbook = openpyxl.Workbook()

        # Remove default 'Sheet' created by openpyxl
        if 'Sheet' in cleaned_workbook.sheetnames:
            cleaned_workbook.remove(cleaned_workbook['Sheet'])
        if 'Sheet' in excluded_workbook.sheetnames:
            excluded_workbook.remove(excluded_workbook['Sheet'])

        # Create the primary sheets for processed data, inserting at the beginning
        cleaned_ws = cleaned_workbook.create_sheet(title=output_sheet_name, index=0)
        excluded_ws = excluded_workbook.create_sheet(title=output_sheet_name, index=0)
        cleaned_ws.sheet_state = 'visible'
        excluded_ws.sheet_state = 'visible'

        # --- EXPLICITLY COPY HEADER ROW TO BOTH CLEANED AND EXCLUDED SHEETS (Skipping UUID) ---
        current_row_in_cleaned = cleaned_ws.max_row + 1 # Should be 1
        current_row_in_excluded = excluded_ws.max_row + 1 # Should be 1

        target_col_idx_cleaned = 0 # Counter for column index in the new sheet
        target_col_idx_excluded = 0 # Counter for column index in the new sheet

        for c_idx, original_cell in enumerate(header_row_cells):
            if c_idx == uuid_col_index: # Skip the UUID column
                continue

            # Copy to cleaned sheet
            new_cleaned_cell = cleaned_ws.cell(row=current_row_in_cleaned, column=target_col_idx_cleaned + 1, value=original_cell.value)
            if original_cell.font: new_cleaned_cell.font = original_cell.font.copy()
            if original_cell.fill: new_cleaned_cell.fill = original_cell.fill.copy()
            if original_cell.border: new_cleaned_cell.border = original_cell.border.copy()
            if original_cell.alignment: new_cleaned_cell.alignment = original_cell.alignment.copy()
            new_cleaned_cell.number_format = original_cell.number_format
            if original_cell.protection: new_cleaned_cell.protection = original_cell.protection.copy()
            if original_cell.hyperlink:
                new_cleaned_cell.hyperlink = Hyperlink(ref=original_cell.hyperlink.ref,
                                                       target=original_cell.hyperlink.target,
                                                       tooltip=original_cell.hyperlink.tooltip,
                                                       display=original_cell.hyperlink.display)
            target_col_idx_cleaned += 1 # Increment only if cell was copied

            # Copy to excluded sheet
            new_excluded_cell = excluded_ws.cell(row=current_row_in_excluded, column=target_col_idx_excluded + 1, value=original_cell.value)
            if original_cell.font: new_excluded_cell.font = original_cell.font.copy()
            if original_cell.fill: new_excluded_cell.fill = original_cell.fill.copy()
            if original_cell.border: new_excluded_cell.border = original_cell.border.copy()
            if original_cell.alignment: new_excluded_cell.alignment = original_cell.alignment.copy()
            new_excluded_cell.number_format = original_cell.number_format
            if original_cell.protection: new_excluded_cell.protection = original_cell.protection.copy()
            if original_cell.hyperlink:
                new_excluded_cell.hyperlink = Hyperlink(ref=original_cell.hyperlink.ref,
                                                       target=original_cell.hyperlink.target,
                                                       tooltip=original_cell.hyperlink.tooltip,
                                                       display=original_cell.hyperlink.display)
            target_col_idx_excluded += 1 # Increment only if cell was copied

        # Copy row dimension for the header row
        original_header_row_dim = input_sheet.row_dimensions[1]
        cleaned_ws.row_dimensions[current_row_in_cleaned].height = original_header_row_dim.height
        excluded_ws.row_dimensions[current_row_in_excluded].height = original_header_row_dim.height

        # Foreign character pattern (allowing emojis and common symbols)
        foreign_character_pattern = r'[\u4E00-\u9FFF\uAC00-\uD7AF\u0900-\u097F\u0600-\u06FF\u0400-\u04FF]'

        # Iterate through DATA ROWS of the input sheet (starting from row 2, skipping the header)
        for r_idx, row_cells in enumerate(input_sheet.iter_rows(min_row=2), 2): # Start from 2 to get actual row number
            # Get the cell from the 'KONTEN' column for the current row
            # Ensure it's within bounds for the current row_cells list
            if konten_col_index >= len(row_cells):
                # If the row is shorter than expected and doesn't have a KONTEN column, skip or handle
                continue

            konten_cell = row_cells[konten_col_index]
            # Get the value from the cell, convert to string for processing, handle None values
            konten_value = str(konten_cell.value or '').strip()

            # Determine if this row should be excluded
            exclude_row = False

            # Check for keywords exclusion (case-insensitive)
            for keyword in keywords_list:
                if str(keyword).strip().lower() in konten_value.lower():
                    exclude_row = True
                    break # Found a keyword, exclude this row

            # Check for foreign language characters (only if not already excluded by keyword)
            if not exclude_row:
                if re.search(foreign_character_pattern, konten_value):
                    exclude_row = True

            # Decide which sheet(s) to copy this row to
            target_sheets = []
            if exclude_row:
                target_sheets.append(excluded_ws)
            else:
                target_sheets.append(cleaned_ws)

            # Copy cells to the target sheets, preserving all properties
            for target_ws in target_sheets:
                current_row_in_target = target_ws.max_row + 1
                target_col_idx_data = 0 # Reset column index for data rows
                for c_idx, original_cell in enumerate(row_cells):
                    if c_idx == uuid_col_index: # Skip the UUID column
                        continue
                    
                    new_cell = target_ws.cell(row=current_row_in_target, column=target_col_idx_data + 1, value=original_cell.value)

                    # Copy style properties
                    if original_cell.font: new_cell.font = original_cell.font.copy()
                    if original_cell.fill: new_cell.fill = original_cell.fill.copy()
                    if original_cell.border: new_cell.border = original_cell.border.copy()
                    if original_cell.alignment: new_cell.alignment = original_cell.alignment.copy()
                    new_cell.number_format = original_cell.number_format
                    if original_cell.protection: new_cell.protection = original_cell.protection.copy()

                    # Explicitly copy hyperlink property
                    if original_cell.hyperlink:
                        new_cell.hyperlink = Hyperlink(ref=original_cell.hyperlink.ref,
                                                       target=original_cell.hyperlink.target,
                                                       tooltip=original_cell.hyperlink.tooltip,
                                                       display=original_cell.hyperlink.display)
                    target_col_idx_data += 1 # Increment only if cell was copied

            # Copy row dimensions (height) from original to target sheets for data rows
            original_data_row_dim = input_sheet.row_dimensions[r_idx]
            for target_ws in target_sheets:
                target_ws.row_dimensions[current_row_in_target].height = original_data_row_dim.height

        # Copy column dimensions (width) from input sheet to the primary processed sheets
        # This needs to be done after all rows are written to ensure it's applied correctly
        for col_dim in input_sheet.column_dimensions.values():
            # Only copy if width is explicitly set and it's not the UUID column
            if col_dim.width and (col_dim.index is None or openpyxl.utils.column_index_from_string(col_dim.index) -1 != uuid_col_index): # Adjust index for 0-based
                cleaned_ws.column_dimensions[col_dim.index].width = col_dim.width
                excluded_ws.column_dimensions[col_dim.index].width = col_dim.width

        # Copy all other sheets from the original workbook to both output workbooks
        for workbook in [cleaned_workbook, excluded_workbook]:
            for sheet_name in original_workbook.sheetnames:
                # Do not copy the input sheet itself, as its data is handled by the row-by-row logic above
                if sheet_name != input_sheet_name:
                    original_sheet = original_workbook[sheet_name]
                    # Create a new sheet in the output workbook with the same title
                    new_ws = workbook.create_sheet(title=sheet_name)
                    new_ws.sheet_state = original_sheet.sheet_state # Preserve hidden/visible state

                    # Copy all cells, including formatting and hyperlinks
                    target_other_sheet_col_idx = 0 # Counter for columns in other sheets
                    for row in original_sheet.iter_rows(values_only=False):
                        for c_idx, cell in enumerate(row):
                            if c_idx == uuid_col_index: # Skip the UUID column even in other sheets if present
                                continue
                            
                            new_cell = new_ws.cell(row=cell.row, column=target_other_sheet_col_idx + 1, value=cell.value) # Use cell.row to maintain original row position for other sheets

                            if cell.font: new_cell.font = cell.font.copy()
                            if cell.fill: new_cell.fill = cell.fill.copy()
                            if cell.border: new_cell.border = cell.border.copy()
                            if cell.alignment: new_cell.alignment = cell.alignment.copy()
                            new_cell.number_format = cell.number_format
                            if cell.protection: new_cell.protection = cell.protection.copy()
                            if cell.hyperlink:
                                new_cell.hyperlink = Hyperlink(ref=cell.hyperlink.ref,
                                                               target=cell.hyperlink.target,
                                                               tooltip=cell.hyperlink.tooltip,
                                                               display=cell.hyperlink.display)
                            target_other_sheet_col_idx += 1 # Increment only if cell was copied

                    # Copy column dimensions (and skip UUID if identified)
                    for col_dim in original_sheet.column_dimensions.values():
                        if col_dim.width and (col_dim.index is None or openpyxl.utils.column_index_from_string(col_dim.index) -1 != uuid_col_index):
                            new_ws.column_dimensions[col_dim.index].width = col_dim.width
                    # Copy row dimensions
                    for row_dim in original_sheet.row_dimensions.values():
                        if row_dim.height:
                            new_ws.row_dimensions[row_dim.index].height = row_dim.height


        # Auto-fit column widths for the main processed sheets for better readability
        # This will only apply to the sheets with new data, not the copied raw sheets.
        for ws in [cleaned_ws, excluded_ws]:
            for column_idx, column in enumerate(ws.columns):
                # Ensure we only process existing columns and apply auto-fit based on actual content
                max_length = 0
                for cell in column:
                    try:
                        if cell.value is not None:
                            current_length = len(str(cell.value))
                            if current_length > max_length:
                                max_length = current_length
                    except TypeError:
                        pass
                # Get the column letter for the current column_idx
                column_letter = openpyxl.utils.get_column_letter(column_idx + 1)
                adjusted_width = min((max_length + 2), 75)
                if adjusted_width > 0:
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbooks
        cleaned_workbook.save(cleaned_output_filepath)
        excluded_workbook.save(excluded_output_filepath)

        print(f"Cleaned data saved to {cleaned_output_filepath} in sheet '{output_sheet_name}' and other sheets preserved.")
        print(f"Excluded items saved to {excluded_output_filepath} in sheet '{output_sheet_name}' and other sheets preserved.")

    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found at {input_filepath}")
    except ValueError as e:
        raise e
    except Exception as e:
        # Catch any other unexpected errors during openpyxl operations
        raise Exception(f"An error occurred during data processing: {e}")
    finally:
        # Delete the original uploaded input file and the temporary file (if created)
        # Add small delay to ensure file handles are released before deletion
        time.sleep(0.1) # Small delay to help ensure file handles are released
        if os.path.exists(input_filepath):
            try:
                os.remove(input_filepath)
                print(f"Deleted uploaded input file: {input_filepath}")
            except Exception as e:
                print(f"Error deleting input file {input_filepath}: {e}")
        # The temporary file is no longer created by the new openpyxl-direct processing logic
        # if os.path.exists(temp_input_filepath):
        #     try:
        #         os.remove(temp_input_filepath)
        #         print(f"Deleted temporary working file: {temp_input_filepath}")
        #     except Exception as e:
        #         print(f"Error deleting temporary file {temp_input_filepath}: {e}")
